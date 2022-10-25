from genericpath import exists, isdir
import os
from posixpath import expanduser
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook


class Sheet:

    def __init__(self):
        self.TITLES = ["Dia", "Inicio1", "Fin1",
                       "Inicio2", "Fin2", "Turno", "Total"]
        self.workbook = Workbook()
        self.workbook.iso_dates = True
        self.worksheet = self.workbook.active

    def setStyles(self):
        """ Area necesaria:
            a4:g12
            Color #595959
        """
        thin = Side(border_style="thin", color="000000")
        # worksheet.merge_cells(row.INITIAL_CELL+':B5')
        self.worksheet.merge_cells("A4"+':B5')
        # initial_cell = worksheet[row.INITIAL_CELL]
        # Set grid

        # Set bg_color
        initial_cell = self.worksheet["A4"]
        initial_cell.border = Border(
            top=thin, left=thin, right=thin, bottom=thin)

    def saveFile(self, week, user):
        folder_name = "kw" + str(week)
        folder_path = os.getcwd()
        file_name = str.capitalize(user)+"-"+str(week)+".xlsx"
        file_relative_path = folder_name+"\\"+file_name
        path_to_file = file_relative_path+"\\" + file_name
        if not isdir(folder_name):
            os.mkdir(folder_name)
            print("created folder : ", folder_name)
        # Save the file
        if not exists(path_to_file):
            self.workbook.save(file_relative_path)
            expanduser(folder_path+"\\"+file_relative_path)
            print("created file : ", file_name)
        else:
            print("file exists")

    def setHeader(self, week, row):
        logo = Image('res\\logo.png')
        self.worksheet.add_image(logo, row.get_position())  # size A4:B5
        row.next_row(2)
        self.worksheet['C'+str(row.n_row)] = "kw"
        self.worksheet['D'+str(row.n_row)] = week
        self.worksheet.append(self.TITLES)
        row.next_row()

    def setContent(self, days, row):
        for i in range(len(days)):
            day = days[i]
            formula = row.get_formula_total()
            new_row = [day, "", "", "",
                            "", "", "="+formula]
            self.worksheet.append(new_row)
            row.next_row()


""" 
def workbookConfig():

    wb
    ws = wb.active  # grab the active worksheet
    setStyles(ws)
    return ws """
